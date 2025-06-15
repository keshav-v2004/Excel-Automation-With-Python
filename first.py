import openpyxl as xl

wb = xl.load_workbook('transactions.xlsx') #load the entire excel file and return a sheet object
sheet = wb['Sheet1'] # load the first sheet

cell = sheet["A1"] # value of cell gets loaded
print(cell.value)
print()

print(sheet.max_row) ## returns the maximum number of rows in the sheet (also includes the header i.e row names)

newColumnToBeAdded = sheet.cell(1,4)
newColumnToBeAdded.value = "Updated prices" # new column name added

print()
for eachRow in range(2,sheet.max_row+1):

    oldCellValue = sheet.cell(eachRow,3).value

    newCellValue = oldCellValue * 0.9 ## this new value will be added in a new column correspondingly

    newColumn = sheet.cell(eachRow,4) ## get the location of column where updated prices will be stored
    newColumn.value = newCellValue ## assigning value to the cell correspondingly


wb.save('transactions.xlsx') ## save the changes <file-name-given-as-argument>